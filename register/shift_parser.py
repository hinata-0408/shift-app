import re
import calendar
from datetime import date
from typing import Optional, Dict, List
import openpyxl


def parse_shift_code(code) -> Optional[Dict]:
    """シフトコード文字列を出退勤情報に変換する。休みの場合はNoneを返す。"""
    if code is None:
        return None

    code = str(code).strip()

    if code == '' or code == 'None':
        return None

    if code in ['有給', '公休', '休', '代休']:
        return None

    # 会議・応援は出勤扱いだがレジ免除
    if code in ['会', '応']:
        return {
            'start_hour': 9, 'start_min': 30,
            'end_hour': 19, 'end_min': 0,
            'shift_type': code,
            'is_register_duty': False,
        }

    if code == '早':
        return {'start_hour': 9, 'start_min': 30, 'end_hour': 19, 'end_min': 0, 'shift_type': '早', 'is_register_duty': True}

    if code == '遅':
        return {'start_hour': 11, 'start_min': 0, 'end_hour': 20, 'end_min': 30, 'shift_type': '遅', 'is_register_duty': True}

    if code == '中1':
        return {'start_hour': 10, 'start_min': 0, 'end_hour': 19, 'end_min': 30, 'shift_type': '中', 'is_register_duty': True}

    if code == '中2':
        return {'start_hour': 10, 'start_min': 30, 'end_hour': 20, 'end_min': 0, 'shift_type': '中', 'is_register_duty': True}

    if code == '流早':
        return {'start_hour': 9, 'start_min': 15, 'end_hour': 18, 'end_min': 45, 'shift_type': '流早', 'is_register_duty': True}

    if code == '研':
        return {'start_hour': 9, 'start_min': 30, 'end_hour': 18, 'end_min': 0, 'shift_type': '研', 'is_register_duty': False}

    if code == '初売':
        return {'start_hour': 9, 'start_min': 30, 'end_hour': 19, 'end_min': 0, 'shift_type': '早', 'is_register_duty': True}

    # 流XX: 9:15開始、XX:00終了
    m = re.match(r'^流(\d{1,2})$', code)
    if m:
        return {'start_hour': 9, 'start_min': 15, 'end_hour': int(m.group(1)), 'end_min': 0, 'shift_type': '流早', 'is_register_duty': True}

    # 流XX3: 9:15開始、XX:30終了
    m = re.match(r'^流(\d{1,2})3$', code)
    if m:
        return {'start_hour': 9, 'start_min': 15, 'end_hour': int(m.group(1)), 'end_min': 30, 'shift_type': '流早', 'is_register_duty': True}

    # 早XX / 早XX3
    m = re.match(r'^早(\d{1,2})(3?)$', code)
    if m:
        return {'start_hour': 9, 'start_min': 30, 'end_hour': int(m.group(1)), 'end_min': 30 if m.group(2) else 0, 'shift_type': '早', 'is_register_duty': True}

    # XX遅 / XX3遅
    m = re.match(r'^(\d{1,2})(3?)遅$', code)
    if m:
        return {'start_hour': int(m.group(1)), 'start_min': 30 if m.group(2) else 0, 'end_hour': 20, 'end_min': 30, 'shift_type': '遅', 'is_register_duty': True}

    # XX3YY3 (例: 103193 = 10:30〜19:30)
    m = re.match(r'^(\d{2})3(\d{2})3$', code)
    if m:
        sh = int(m.group(1))
        return {'start_hour': sh, 'start_min': 30, 'end_hour': int(m.group(2)), 'end_min': 30, 'shift_type': '早' if sh < 11 else '遅', 'is_register_duty': True}

    # XX3YY (例: 10319 = 10:30〜19:00)
    m = re.match(r'^(\d{2})3(\d{2})$', code)
    if m:
        sh = int(m.group(1))
        return {'start_hour': sh, 'start_min': 30, 'end_hour': int(m.group(2)), 'end_min': 0, 'shift_type': '早' if sh < 11 else '遅', 'is_register_duty': True}

    # XXYY3 (例: 12193 = 12:00〜19:30)
    m = re.match(r'^(\d{2})(\d{2})3$', code)
    if m:
        sh = int(m.group(1))
        return {'start_hour': sh, 'start_min': 0, 'end_hour': int(m.group(2)), 'end_min': 30, 'shift_type': '早' if sh < 11 else '遅', 'is_register_duty': True}

    # XXYY (例: 1119 = 11:00〜19:00)
    m = re.match(r'^(\d{2})(\d{2})$', code)
    if m:
        sh = int(m.group(1))
        return {'start_hour': sh, 'start_min': 0, 'end_hour': int(m.group(2)), 'end_min': 0, 'shift_type': '早' if sh < 11 else '遅', 'is_register_duty': True}

    print(f"  [警告] 認識できないシフトコード: {code}")
    return None


def parse_shift_excel(file_path: str, target_year: int, target_month: int):
    """シフト表Excelを解析し、(日付→出勤情報リスト, 全スタッフリスト)を返す。"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['3）入力用']

    result = {}
    all_staff_list = []

    # リーダー行を起点に各部門の範囲を特定する
    leader_rows = []
    for row in range(1, 420):
        role = ws.cell(row=row, column=5).value
        if role and ('リーダー' in str(role) or 'ﾘｰﾀﾞｰ' in str(role)):
            leader_rows.append(row)

    dept_names = ['季節', '家電', '情報', '通信']
    dept_ranges = []

    if leader_rows:
        dept_ranges.append({'dept': '店長', 'start_row': 34, 'end_row': leader_rows[0] - 1})

    for i, leader_row in enumerate(leader_rows):
        dept_name = dept_names[i] if i < len(dept_names) else f'部門{i+1}'
        end_row = leader_rows[i + 1] - 1 if i + 1 < len(leader_rows) else 400
        dept_ranges.append({'dept': dept_name, 'start_row': leader_row, 'end_row': end_row})

    days_in_month = calendar.monthrange(target_year, target_month)[1]

    for dept_info in dept_ranges:
        dept = dept_info['dept']
        for row in range(dept_info['start_row'], dept_info['end_row'] + 1):
            cd = ws.cell(row=row, column=4).value
            name = ws.cell(row=row, column=6).value

            if not cd or not name:
                continue
            try:
                int(cd)
            except Exception:
                continue

            all_staff_list.append({'name': str(name).strip(), 'department': dept, 'row_number': row})

            for day in range(1, days_in_month + 1):
                shift_code = ws.cell(row=row, column=9 + day).value
                shift_data = parse_shift_code(shift_code)
                if shift_data:
                    target_date = date(target_year, target_month, day)
                    if target_date not in result:
                        result[target_date] = []
                    result[target_date].append({
                        'name': str(name).strip(),
                        'department': dept,
                        'start_hour': shift_data['start_hour'],
                        'start_min': shift_data['start_min'],
                        'end_hour': shift_data['end_hour'],
                        'end_min': shift_data['end_min'],
                        'shift_type': shift_data['shift_type'],
                        'is_register_duty': shift_data['is_register_duty'],
                        'row_number': row,
                    })

    wb.close()
    return result, all_staff_list


def import_shift_to_db(file_path: str, target_year: int, target_month: int):
    """Excelシフト表をDBにインポートする。既存スタッフのレベルは維持し、新規はLv0で登録。"""
    from .models import Staff, DailyAttendance

    shift_data, all_staff_list = parse_shift_excel(file_path, target_year, target_month)

    created_count = 0
    updated_count = 0
    staff_map = {}

    for staff_info in all_staff_list:
        name = staff_info['name']
        dept = staff_info['department']
        row_number = staff_info['row_number']

        existing = Staff.objects.filter(name=name).first()
        if existing:
            existing.department = dept
            existing.display_order = row_number
            existing.is_active = True
            existing.save()
            staff_map[name] = existing
            updated_count += 1
        else:
            new_staff = Staff.objects.create(
                name=name, department=dept, level=0,
                display_order=row_number, is_active=True,
            )
            staff_map[name] = new_staff
            created_count += 1

    days_in_month = calendar.monthrange(target_year, target_month)[1]
    DailyAttendance.objects.filter(
        date__gte=date(target_year, target_month, 1),
        date__lte=date(target_year, target_month, days_in_month),
    ).delete()

    imported_count = 0
    for target_date, attendances in shift_data.items():
        for att in attendances:
            staff = staff_map.get(att['name'])
            if staff:
                DailyAttendance.objects.create(
                    date=target_date, staff=staff,
                    start_hour=att['start_hour'], start_min=att['start_min'],
                    end_hour=att['end_hour'], end_min=att['end_min'],
                    shift_type=att['shift_type'],
                )
                imported_count += 1

    return {
        'imported': imported_count,
        'staff_created': created_count,
        'staff_updated': updated_count,
        'dates': len(shift_data),
        'total_staff': len(all_staff_list),
    }
